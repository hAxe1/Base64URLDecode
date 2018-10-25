#!/usr/bin/python
import base64
import re
import openpyxl
import urllib 
import itertools
import sys
from os.path import dirname, abspath


if len(sys.argv) == 2:
    fname = sys.argv[1]
    fnamelwr = fname.lower() 
    resfname = "Base64Decoded.xlsx"

elif len(sys.argv) == 3:
    fname = sys.argv[1]
    fnamelwr = fname.lower() 
    resfname = sys.argv[2]
    resfnamelwr = resfname.lower()

    if not resfnamelwr.endswith(".xlsx"):
        print("Results file must be an XLSX")
        sys.exit(1)

else:
    print("Incorrect Syntax, use: " + __file__ +" filename.xlsx \nYou may optionally specify the filename for output file: " + __file__ + " filename.xlsx output.xlsx")
    sys.exit(1)

print('Loading list...')
xl_sheet = openpyxl.Workbook().active #Workbook to read data from

if fnamelwr.endswith(".csv"): #If CSV given, convert to xlsx for script
    import csv
    print('Converting CSV to XLSX')
    with open(fname, 'rb') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader, start=1):
            for c, val in enumerate(row, start=1):
                xl_sheet.cell(row=r, column=c).value = val

elif fnamelwr.endswith(".xls"): #If XLS given, convert to XLSX
    import xlrd
    xlsBook = xlrd.open_workbook(fname)
    xlsSheet = xlsBook.sheet_by_index(0)
    print('Converting XLS to XLSX')
    for row in xrange(0, xlsSheet.nrows):
        for col in xrange(0, xlsSheet.ncols):
            xl_sheet.cell(row=row+1, column=col+1).value = xlsSheet.cell_value(row, col)

elif fnamelwr.endswith(".xlsx"): #If XLSX given
    xl_sheet = openpyxl.load_workbook(fname, read_only=True).active

else:
    print("File must be a CSV, XLS, or  XLSX")
    sys.exit(1)

spinner = itertools.cycle(['-', '/', '|', '\\']) #Spinning Cursor
prevb64 = set() #Set for keeping track of all Base64 strings found
#Start Builiding XLSs
wb = openpyxl.Workbook() #Workbook to save output to
ws = wb.create_sheet("Base64", 0)
newsheetrow_idx = 2 #Index to track rows on output spreadsheet
print("Processing list...")

ws.cell(row=1, column=1, value="URL")
ws.cell(row=1, column=2, value="Base64 String")
ws.cell(row=1, column=3, value="Decoded Value")


for row_idx in range(1, xl_sheet.max_row):   #Iterate through spreadsheet 
    sys.stdout.write(spinner.next())  # Spinning cursor
    sys.stdout.flush()                
    sys.stdout.write('\b') 

    cell_obj = xl_sheet.cell(row=row_idx, column=1)  # Get cell object by row, column
    urlval = cell_obj.value #Variable containing URL from sheet
    #Regex looking for data that looks like base64, looking for 8 or more characters to weed out some of the noise, may increase the required length
    b64str = re.match(r'(?:^.*\/.*[^?&#=]=)((?:[A-Za-z0-9+%]{4}){2,}(?:(?:[A-Za-z0-9+%]{3}(?:=|%3D))|(?:[A-Za-z0-9+%]{2}(?:=|%3D){2}))?)(?:&.*|$)',urlval)
    if b64str: 
        b64str = b64str.group(1)
        urldecodedb64 = urllib.unquote(b64str) #URL Decode the output from regex incase it was URL encoded
        if urldecodedb64 not in prevb64: #Check to see if Base64 String has been found in this sheet already, if so, skip the URL
            try:
                if urldecodedb64 != " ": #May turn into a list/set containing all values I want to skip
                    b64strdecode = base64.b64decode(urldecodedb64)
            except:
                continue #Skip to next Base64 string if decoding failed     
            for letter in b64strdecode:    #Loop through the characters of the decoded string to ensure they are ASCII characters
                if ord(letter) < 32 or ord(letter) > 127:
                    break #If non-ASCII character found, move to next Base64 string
            else: #Otherwise, write the URL, Base64 string, and Decoded Base64 string to the sheet
                ws.cell(row=newsheetrow_idx, column=1, value=urlval)
                ws.cell(row=newsheetrow_idx, column=2, value=urldecodedb64)
                ws.cell(row=newsheetrow_idx, column=3, value=b64strdecode)
                newsheetrow_idx+=1
                prevb64.add(urldecodedb64) #Add Base64 value to the set

wb.save(resfname) #Save the new output
print('Processing complete, Results saved to ' + dirname(abspath(__file__)) + "/" + resfname)
